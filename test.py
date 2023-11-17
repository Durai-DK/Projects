# from textblob import TextBlob
#
# text = "my job is not a bad or bad Job"
#
# blob = TextBlob(text)
#
# sentiment = blob.sentiment.polarity
#
# if sentiment > 0:
#     print("Positive Sentiment")
#
# elif sentiment < 0:
#     print("Negative Sentiment")
#
# else:
#     print("Neutral Sentiment")

# _____________________________________________________________

# from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
#
# text = "am doesn't bore my Job"
# # text = "am going to job"
#
# analyzer = SentimentIntensityAnalyzer()
#
# sentiment = analyzer.polarity_scores(text)
#
#
# if sentiment["compound"] > 0:
#     print("Positive Sentiment")
#
# elif sentiment["compound"] < 0:
#     print("Negative Sentiment")
#
# else:
#     print("Neutral Sentiment")

# ---------------------------------------------------------------
# import nltk
# nltk.download()
from openpyxl import load_workbook, Workbook
from nltk.sentiment import SentimentIntensityAnalyzer

wb = Workbook()
ws = wb.active

l_wb = load_workbook(r"C:\Users\admin\Documents\reviews.xlsx")
l_ws = l_wb.active

# text = "am bore my Job"
text = "am going to job"

ws['a1'] = "Reviews"
ws['b1'] = "Status"

for r in range(2, l_ws.max_row+1):

    ws.cell(row=r, column=1).value = l_ws.cell(row=r, column=1).value

    text = l_ws.cell(row=r, column=1).value

    analyzer = SentimentIntensityAnalyzer()

    sentiment = analyzer.polarity_scores(text)

    if sentiment["compound"] > 0:
        status = "Positive Sentiment"
        print(status)
        ws.cell(row=r, column=2).value = status


    elif sentiment["compound"] < 0:
        status = "Negative Sentiment"
        print(status)
        ws.cell(row=r, column=2).value = status

    else:
        status = "Neutral Sentiment"
        print(status)
        ws.cell(row=r, column=2).value = status

    wb.save(r"E:\Durai\Projects\Reviews Status.xlsx")